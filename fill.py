version = 1.00

from tendo import singleton
me = singleton.SingleInstance()

import os
import sys
script_path = os.path.dirname(sys.executable)
file_path = os.path.abspath(__file__)

import requests
import re

url = 'https://github.com/Maxwellaf/RSC/blob/main/version.ini'
response = requests.get(url)
with open(r'C:\Users\sd55\Desktop\text.txt', 'wb') as f:
    f.write(response.content)

with open(r'C:\Users\sd55\Desktop\text.txt', 'r') as file:
    # Считываем содержимое файла
    file_content = file.read()
matches = re.search(r'\"blob\":\{\"rawLines\":\[\"vers = (.*)\",\"updateExtra = (.*)\",\"updateMain = (.*)\"\]', file_content)
new_version = matches.group(1)
updateExtra = matches.group(2)
updateMain = matches.group(3)
#"blob":{"rawLines":["vers = 1.00","updateExtra = True","updateMain = True"]
if float(new_version) > float(version) and bool(updateExtra):
    print('Качай новую версию')
    response = requests.get(url)
    with open(r'C:\Users\sd55\Desktop\text.txt', 'wb') as f:
        f.write(response.content)
else:
    print('Не качай новую версию')

from docxtpl import DocxTemplate
import openpyxl
import re
from tkinter import *
from tkinter import messagebox
import tkinter
from datetime import date
import configparser
import os.path
from tkinter import filedialog
from openpyxl import Workbook
config2_path = script_path + "\config.ini"

config = configparser.ConfigParser()
config2 = configparser.ConfigParser()
config2.read(config2_path, encoding="windows-1251")

isFile = False

temp = re.match(r"20(\d+)-(\d+)-(\d+)", str(date.today()))
day = temp.group(3)
month = temp.group(2)
year = temp.group(1)
result_path = None

# import subprocess #библиотека для открытия файлов
    
def saveResult():
    global result_path
    result_path = filedialog.askdirectory()
    if result_path == None:
        return
    
def choosePath():
    global file_path
    file_path = filedialog.askopenfile()
    if file_path == None:
        return
    config.read(file_path.name, encoding="windows-1251")
    isFile = os.path.isfile(file_path.name)
    if isFile:
        update()
        
def save():
    try:
        file_path.name
    except:
        saveAs()
    else:
        with open(file_path.name, 'w') as configfile:
            writeConfig(file_path.name)
        messagebox.showinfo("Успешно", "Файл сохранен")
        
def saveAs():
    if orgType.get() == 1:
        short = "ИПКФХ"
        if len(entryOwner.get()) > 2:
            temp = re.match(r"(\w+) (\w)\w+ (\w)\w+", entryOwner.get())
            short = '№' + entryAppNumber.get() + ' ИПКФХ ' + temp.group(1) + " " + temp.group(2) + "." + temp.group(3) + "."
    elif orgType.get() == 2:
        short = "ООО"
        if len(entryOrg.get()) > 2:
            short += ' ' + entryOrg.get()
            
    save_directory = filedialog.asksaveasfilename(
        defaultextension='.ini', filetypes=[("ini files", '*.ini')],
        initialfile = short,
        title="Выберите директорию и название")
    if save_directory == '':
        return
    writeConfig(save_directory)

def writeConfig(directory):
    config = configparser.ConfigParser()
    config.add_section('settings')
    config.set('settings', 'тип_организации', str(orgType.get()))
    config.set('settings', 'имя_заявителя', entryOwner.get())
    config.set('settings', 'имя_ответственного_лица', entryResponsiblePerson.get())
    config.set('settings', 'основание_полномочий_ответственного_лица', entryResponsiblePersonBase.get())
    config.set('settings', 'дата_основания_полномочий_ответственного_лица', entryResponsiblePersonBaseDate.get())
    config.set('settings', 'имя_заявителя_рп', entryOwnerInclined.get())
    config.set('settings', 'название_оргазизации', entryOrg.get())
    config.set('settings', 'должность_заявителя', entryPost.get())
    config.set('settings', 'основание_полномочий', entryBase.get())
    config.set('settings', 'юридический_адрес', entryAdressLegal.get())
    config.set('settings', 'фактический_адрес', entryAdressFact.get())
    config.set('settings', 'инн', entryInn.get())
    config.set('settings', 'кпп', entryKpp.get())
    config.set('settings', 'огрн', entryOgrn.get())
    config.set('settings', 'огрнип', entryOgrnip.get())
    config.set('settings', 'название_банка', entryBank.get())
    config.set('settings', 'расчетный_счет', entryAccPay.get())
    config.set('settings', 'корреспондентский_счет', entryAccCorr.get())
    config.set('settings', 'бик', entryBik.get())
    config.set('settings', 'номер_телефона', entryNumber.get())
    config.set('settings', 'почта', entryMailOrg.get())
    config.set('settings', 'место_отбора', entrySelectionPlace.get())
    config.set('settings', 'номер_договора', entryContract.get())
    config.set('settings', 'изготовитель', entryMaker.get())
    config.set('settings', 'юридический_адрес_изготовителя', entryMakerAddressLegal.get())
    config.set('settings', 'фактический_адрес_изготовителя', entryMakerAddressFact.get())
    config.set('settings', 'культура', str(culture.get()))
    config.set('settings', 'номер_заявления', entryAppNumber.get())
    config.set('settings', 'номер_договора', entryContract.get())
    config.set('settings', 'масса_партии', entryWeight.get())
    config.set('settings', 'консультационные_услуги', str(service.get()))
    with open(directory, 'w') as configfile:
        config.write(configfile)

def make_menu(w):
    global the_menu
    the_menu = tkinter.Menu(w, tearoff=0)
    the_menu.add_command(label="Вырезать")
    the_menu.add_command(label="Копировать")
    the_menu.add_command(label="Вставить")

def show_menu(e):
    w = e.widget
    the_menu.entryconfigure("Вырезать",
    command=lambda: w.event_generate("<<Cut>>"))
    the_menu.entryconfigure("Копировать",
    command=lambda: w.event_generate("<<Copy>>"))
    the_menu.entryconfigure("Вставить",
    command=lambda: w.event_generate("<<Paste>>"))
    the_menu.tk.call("tk_popup", the_menu, e.x_root, e.y_root)

isRoot2Open = False
priceSurvices = 1500
itogo = 0

def monthName(a):
    if a==1:
        return "января"
    elif a==2:
        return "февраля"
    elif a==3:
        return "марта"
    elif a==4:
        return "апреля"
    elif a==5:
        return "мая" 
    elif a==6:
        return "июня"
    elif a==7:
        return "июля"
    elif a==8:
        return "августа"
    elif a==9:
        return "сентября"
    elif a==10:
        return "октября"
    elif a==11:
        return "ноября"
    elif a==12:
        return "декабря"
        
def about():
    messagebox.showinfo("Информация о программе", 'Данная программа была разработана специально для организации "Россельхозцентр" по Ростовской области')
    
def setOrgType():
    if orgType.get() == 1:
        entryOgrn['state'] = 'disabled'
        entryKpp['state'] = 'disabled'
        entryOgrnip['state'] = 'normal'
        entryOrg['state'] = 'disabled'
        entryPost['state'] = 'disabled'
    if orgType.get() == 2:
        entryOgrn['state'] = 'normal'
        entryOgrnip['state'] = 'disabled'
        entryKpp['state'] = 'normal'
        entryOrg['state'] = 'normal'
        entryPost['state'] = 'normal'

def saveConfig(directory):
    config2.set('settings', 'доверенность', entryHeadBase.get())
    config2.set('settings', 'фио_руководителя', entryHeadFIO.get())
    config2.set('settings', 'фио_руководителя_рп', entryHeadFIOInclined.get())
    config2.set('settings', 'номер_отдела', entryDepId.get())
    config2.set('settings', 'название_отдела', entryDepNameInclined.get())
    config2.set('settings', 'почта', entryMail.get())
    with open(directory, 'w') as configfile:
        config2.write(configfile)

def changeHeadInfo():
    if entryHeadFIO['state'] == 'disabled':
        buttonChangeHeadInfo['text'] = 'Сохранить'
        entryHeadFIO['state'] = 'normal'
        entryHeadBase['state'] = 'normal'
        entryHeadFIOInclined['state'] = 'normal'
        entryDepId['state'] = 'normal'
        entryDepNameInclined['state'] = 'normal'
        entryMail['state'] = 'normal'
    else:
        saveConfig(config2_path)
        buttonChangeHeadInfo['text'] = 'Изменить информацию о начальнике МРО'
        entryHeadFIO['state'] = 'disabled'
        entryHeadBase['state'] = 'disabled'
        entryHeadFIOInclined['state'] = 'disabled'
        entryDepId['state'] = 'disabled'
        entryDepNameInclined['state'] = 'disabled'
        entryMail['state'] = 'disabled'
def str2(n):
    if get_precision(n) == 0:
        return str(n) + ",00"
    elif get_precision(n) == 1:
        return str(n).replace(".", ",") + "0"
    else:
        return str(n).replace(".", ",")
    
def get_precision(f):
    str_f = str(f)
    if '.' not in str_f:
        return 0
    return len(str_f[str_f.index('.') + 1:])

def calculateItogo():
    if gmo.get() == 1 and physics.get() == 1:
        labelItogo.config(text = "Данный комплекс отсутствует")
        return
    startPrice = 0
    if physics.get() == 1 and (cultureName[culture.get()] == "озимая пшеница" or cultureName[culture.get()] == "ячмень"):
        startPrice = startPrice + wheatExtra
    if service.get() == 0:
        startPrice = startPrice + workComplexPrice[cultureName[culture.get()]][physics.get() + gmo.get()*2]
        nds = round(startPrice*0.2, 2)
    else:
        startPrice = startPrice + workComplexPrice[cultureName[culture.get()]][physics.get() + gmo.get()*2] + priceSurvices
        nds = round(startPrice*0.2, 2)
    itogo = round(nds + startPrice, 2)
    labelItogo.config(text = "Итоговая сумма: " + str2(itogo) + " рублей")
    
def update():
    try:
        updateRadio(orgType, 'тип_организации')
        updateEntry(entryOwner, 'имя_заявителя')
        updateEntry(entryResponsiblePerson, 'имя_ответственного_лица')
        updateEntry(entryResponsiblePersonBase, 'основание_полномочий_ответственного_лица')
        updateEntry(entryResponsiblePersonBaseDate, 'дата_основания_полномочий_ответственного_лица')
        updateEntry(entryOwnerInclined, 'имя_заявителя_рп')
        updateEntry(entryOrg, 'название_оргазизации')
        updateEntry(entryPost, 'должность_заявителя')
        updateEntry(entryBase, 'основание_полномочий')
        updateEntry(entryAdressLegal, 'юридический_адрес')
        updateEntry(entryAdressFact, 'фактический_адрес')
        updateEntry(entryInn, 'инн')
        updateEntry(entryKpp, 'кпп')
        updateEntry(entryOgrn, 'огрн')
        updateEntry(entryOgrnip, 'огрнип')
        updateEntry(entryBank, 'название_банка')
        updateEntry(entryAccPay, 'расчетный_счет')
        updateEntry(entryAccCorr, 'корреспондентский_счет')
        updateEntry(entryBik, 'бик')
        updateEntry(entryNumber, 'номер_телефона')
        updateEntry(entryMailOrg, 'почта')
        updateEntry(entrySelectionPlace, 'место_отбора')
        updateEntry(entryContract, 'номер_договора')
        updateEntry(entryMaker, 'изготовитель')
        updateEntry(entryMakerAddressLegal, 'юридический_адрес_изготовителя')
        updateEntry(entryMakerAddressFact, 'фактический_адрес_изготовителя')
        updateRadio(culture, 'культура')
        updateEntry(entryContract, 'номер_договора')
        updateEntry(entryAppNumber, 'номер_заявления')
        updateEntry(entryWeight, 'масса_партии')
        updateCheck(service, 'консультационные_услуги')
    except:
        messagebox.showinfo("Ошибка", "Файл настроек повреждён")

def updateCheck(check, text):
    try:
        if config['settings'][text] == '1':
            check.set(1)
        else:
            check.set(0)
    except:
        print('Поле ' + text + ' - пустое')

def updateRadio(radio, text):
    try:
        radio.set(int(config['settings'][text]))
    except:
        print('Поле ' + text + ' - пустое')
        radio.set(1)

def updateEntry(entry, text):
    entry.delete(0, END)
    try:
        entry.insert(0, config['settings'][text])
    except:
        print('Поле ' + text + ' - пустое')

def doWeClear():
    result = messagebox.askyesno(title="Подтвержение операции", message="Вы уверены, что хотите очистить поля?")
    if result:
        clear()

def clearEntry(entry):
    entry.delete(0, END)

def clear():
    file_path = None
    clearEntry(entryOwner)
    clearEntry(entryOrg)
    clearEntry(entryPost)
    clearEntry(entryBase)
    clearEntry(entryAdressLegal)
    clearEntry(entryAdressFact)
    clearEntry(entryInn)
    clearEntry(entryKpp)
    clearEntry(entryOgrn)
    clearEntry(entryOgrnip)
    clearEntry(entryBank)
    clearEntry(entryAccPay)
    clearEntry(entryAccCorr)
    clearEntry(entryBik)
    clearEntry(entryNumber)
    clearEntry(entryMailOrg)
    clearEntry(entrySelectionPlace)
    clearEntry(entryContract)
    clearEntry(entryMaker)
    clearEntry(entryMakerAddressLegal)
    clearEntry(entryMakerAddressFact)
    clearEntry(entryOwnerInclined)
    clearEntry(entryResponsiblePerson)
    clearEntry(entryResponsiblePersonBase)
    clearEntry(entryResponsiblePersonBaseDate)
    clearEntry(entryWeight)
    clearEntry(entryCollectYear)
    orgType.set(1)
    culture.set(1)
    service.set(0)

def check():
    if not re.fullmatch(r"\w+ \w+ \w+", entryOwner.get()):
        fail("Некорректно введено ФИО главы (директора)")
        return False
    if not re.fullmatch(r"\w+ \w+ \w+", entryOwnerInclined.get()):
        fail("Некорректно введено ФИО главы (директора) в родительном падеже")
        return False
    if not re.fullmatch(r"\d\d\.\d\d", entryDate.get()):
        fail("Некорректно введена дата")
        return False
    if not re.fullmatch(r"\w+ \w+ \w+", entryResponsiblePerson.get()):
        fail("Некорректно введено ответственное лицо")
        return False
    if not re.fullmatch(r"\w+ .+", entryResponsiblePersonBase.get()):
        fail("Некорректно введен документ, подтверждающий полномочия ответственного лица")
        return False
    if not re.fullmatch(r"\d\d\.\d\d.\d\d\d\d", entryResponsiblePersonBaseDate.get()):
        fail("Некорректно введена дата документа, подтверждающего полномочия ответственного лица")
        return False
    if not re.fullmatch(r"\w+ .+", entryBase.get()):
        fail("Некорректно введен документ, подтверждающий полномочия")
        return False
    if not re.fullmatch(r".+", entryAdressLegal.get()):
        fail("Некорректно введен юридический адрес")
        return False
    if not re.fullmatch(r".+", entryAdressFact.get()):
        fail("Некорректно введен фактический адрес")
        return False
    if not re.fullmatch(r"\d+", entryInn.get()):
        fail("Некорректно введен ИНН")
        return False
    if orgType.get() == 2 and not re.fullmatch(r"\d+", entryKpp.get()):
        fail("Некорректно введен КПП")
        return False
    if orgType.get() == 2 and not re.fullmatch(r"\d+", entryOgrn.get()):
        fail("Некорректно введен ОГРН")
        return False
    if orgType.get() == 1 and not re.fullmatch(r"\d+", entryOgrnip.get()):
        fail("Некорректно введен ОГРНИП")
        return False
    if not re.fullmatch(r".+", entryBank.get()):
        fail("Некорректно введено название банка")
        return False
    if not re.fullmatch(r"\d+", entryAccPay.get()):
        fail("Некорректно введен рассчетный счет")
        return False
    if not re.fullmatch(r"\d+", entryAccCorr.get()):
        fail("Некорректно введен корреспондентский счет")
        return False
    if not re.fullmatch(r"\d+", entryBik.get()):
        fail("Некорректно введен БИК")
        return False
    if not re.fullmatch(r".*\d.*", entryNumber.get()):
        fail("Некорректно введен номер")
        return False
    if not re.fullmatch(r".+@.+", entryMailOrg.get()):
        fail("Некорректно введена почта")
        return False
    if not re.fullmatch(r".+", entrySelectionPlace.get()):
        fail("Некорректно введено место отбора")
        return False
    if not re.fullmatch(r".+", entryWeight.get()):
        entryWeight.insert(0, " ")
    if not re.fullmatch(r".*\d+.*", entryAppNumber.get()):
        fail("Некорректно введен номер заявления")
        return False
    if orgType.get() != 1 and not re.fullmatch(r".+", entryOrg.get()):
        fail("Некорректно введено название")
        return False
    if orgType.get() != 1 and not re.fullmatch(r"\w+.*", entryPost.get()):
        fail("Некорректно введена должность")
        return False
    if gmo.get() == 1 and physics.get() == 1:
        fail("Такого комплекса не существует")
        return False
    if not re.fullmatch(r"\d*", entryContract.get()):
        fail("Некорректно введен номер доовора")
        
    return True

def runApp():
    global result_path
    if not check():
        return False
    saveResult()
    global context
    # calculateItogo() begin
    if gmo.get() == 1 and physics.get() == 1:
        labelItogo.config(text = "Данный комплекс отсутствует")
        return
    startPrice = 0
    if physics.get() == 1 and (cultureName[culture.get()] == "озимая пшеница" or cultureName[culture.get()] == "ячмень"):
        startPrice = startPrice + wheatExtra
    if service.get() == 0:
        startPrice = startPrice + workComplexPrice[cultureName[culture.get()]][physics.get() + gmo.get()*2]
        nds = round(startPrice*0.2, 2)
    else:
        startPrice = startPrice + workComplexPrice[cultureName[culture.get()]][physics.get() + gmo.get()*2] + priceSurvices
        nds = round(startPrice*0.2, 2)
    itogo = round(nds + startPrice, 2)
    
    temp = re.match(r"(\w+) (\w+) (\w+)", entryOwner.get())
    #if morph.parse(temp.group(2))[0].tag.gender == "masc":
    #    gender = "male"
    #else:
    #    gender = "female"
    
    temp = re.match(r"(\w+) (\w)\w+ (\w)\w+", entryOwner.get())
    ownerAbbreviated = temp.group(1) + " " + temp.group(2) + "." + temp.group(3) + "."
    
    temp = re.match(r"(\w+) (\w)\w+ (\w)\w+", entryHeadFIO.get())
    headAbbreviated = temp.group(1) + " " + temp.group(2) + "." + temp.group(3) + "."
    
    temp = re.match(r"(\d+)\.0*(\d+)", entryDate.get())
    monthGroup = re.match(r"\d+\.(\d+)", entryDate.get())
    month = monthGroup.group(1)
    dateText = temp.group(1) + ' ' + monthName(int(temp.group(2)))
    
    temp = re.match(r"(\w+)(.+)", entryBase.get())
    baseInclined = entryBase.get()
    #if temp.group(1).title() == "Устав":
    #    tempWord = "Устава"
    #else:
    #    base = morph.parse(temp.group(1))[0].inflect({'gent'})
    #    tempWord = base.word.title()
    #baseInclined = tempWord + temp.group(2)
    
    temp = re.match(r"(\w+)(.*)", entryPost.get())
    postInclined = "Индивидуального предпринимателя главы крестьянского (фермерского) хозяйства"
    #if orgType.get() == 2 or orgType.get() == 3:
    #    post = morph.parse(temp.group(1))[0].inflect({'gent'})
    #    postInclined = post.word + temp.group(2)
    
    purpose = startPurpose
    if cultureName[culture.get()] == "подсолнечник":
        purpose = purpose + ", " + sunflowerPurpose
    if physics.get():
        purpose = purpose + ", " + purposePhysics[cultureName[culture.get()]]
    
    workComplexNumber = physics.get() + gmo.get()*2
    if (cultureName[culture.get()] == "озимая пшеница" or cultureName[culture.get()] == "ячмень") and physics.get() == 1:
        workComplexNumber = 0
    
    if result_path:
        messagebox.showinfo("Успешно")
    
    context = {
        'date' : entryDate.get(), 
        'month' : month,
        'year' : year,
        'dateText' : dateText,
        'owner' : entryOwner.get(),
        'ownerAbbreviated' : ownerAbbreviated,
        'base' : entryBase.get(),
        'baseInclined' : baseInclined,
        'adressLegal' : entryAdressLegal.get(),
        'adressFact' : entryAdressFact.get(),
        'inn' : entryInn.get(),
        'bank' : entryBank.get(),
        'accPay' : entryAccPay.get(),
        'accCorr' : entryAccCorr.get(),
        'bik' : entryBik.get(),
        'number' : entryNumber.get(),
        'mail' : entryMailOrg.get(),
        'selectionPlace' : entrySelectionPlace.get(),
        'selectionGost' : selectionGost[cultureName[culture.get()]],
        'weight' : entryWeight.get(),
        'appNumber' : entryAppNumber.get(),
        'contract' : entryContract.get(),
        'cultureName' : cultureName[culture.get()],
        'cultureCode' : cultureCode[cultureName[culture.get()]],
        'workComplex' : safety + secondWord[workComplexNumber] + keyWord[cultureName[culture.get()]],
        'workComplexPrice' : str2(workComplexPrice[cultureName[culture.get()]][physics.get() + gmo.get()*2]),
        'priceSurvices' : str2(priceSurvices),
        'startPrice' : str2(startPrice),
        'wheatExtra' : str2(wheatExtra),
        'nds' : str2(nds),
        'itogo' : str2(itogo),
        'purpose' : purpose,
        'headFIO' : entryHeadFIO.get(),
        'headFIOInclined' : entryHeadFIOInclined.get(),
        'headAbbreviated' : headAbbreviated,
        'headBase' : entryHeadBase.get(),
        'depId' : entryDepId.get(),
        'depNameInclined' : entryDepNameInclined.get(),
        'responsiblePerson' : entryResponsiblePerson.get(),
        'responsiblePersonBase' : entryResponsiblePersonBase.get(),
        'responsiblePersonBaseDate' : entryResponsiblePersonBaseDate.get(),
        'mail' : entryMail.get(),
        'collectYear' : entryCollectYeart.get(),
        'maker' : entryMaker.get(),
        'makerAdressLegal' : entryMakerAddressLegal.get(),
        'makerAdressFact' : entryMakerAddressFact.get()
    }
    
    if orgType.get() == 1:
        context['org'] = "Индивидуальный предприниматель глава крестьянского (фермерского) хозяйства " + entryOwner.get()
        context['orgShort'] = "ИП глава К(Ф)Х " + entryOwner.get()
        context['ownerInclined'] = "Индивидуального предпринимателя главы крестьянского (фермерского) хозяйства " + entryOwnerInclined.get()
        context['orgInclined'] = "Индивидуального предпринимателя главы крестьянского (фермерского) хозяйства " + entryOwnerInclined.get()
        context['ogrn_ogrnip_word'] = 'ОГРНИП'
        context['ogrn_ogrnip'] = entryOgrnip.get()
        context['kpp'] = ""
        context['post'] = "ИП глава К(Ф)Х"
    elif orgType.get() == 2:
        context['org'] = "Общество с ограниченной ответственностью " + entryOrg.get()
        context['orgShort'] = "ООО " + entryOrg.get()
        context['ownerInclined'] = postInclined + " " + ownerInclined
        context['orgInclined'] = "Общества с ограниченной ответственностью " + entryOrg.get()
        context['ogrn_ogrnip_word'] = 'ОГРН'
        context['ogrn_ogrnip'] = entryOgrn.get()
        context['kpp'] = "\nКПП: " + entryKpp.get()
        context['post'] = entryPost.get()
    
    if service.get() == 0:
        if physics.get() == 1 and (cultureName[culture.get()] == "озимая пшеница" or cultureName[culture.get()] == "ячмень"):
            doc2 = DocxTemplate(script_path + "\Заявление для пшеницы с ГМО и физ-хим (шаблон).docx")
        else:
            doc2 = DocxTemplate(script_path + "\Заявление (шаблон).docx")
    else:
        if physics.get() == 1 and (cultureName[culture.get()] == "озимая пшеница" or cultureName[culture.get()] == "ячмень"):
            doc2 = DocxTemplate(script_path + "\Заявление для пшеницы с ГМО и физ-хим с услугами (шаблон).docx")
        else:
            doc2 = DocxTemplate(script_path + "\Заявление с услугами (шаблон).docx")
    doc2.render(context)
    if result_path:
        doc2.save(result_path + "\Заявление " + entryAppNumber.get() + ".docx")
    doc3 = DocxTemplate(script_path + "\Справка (шабон).docx")
    doc3.render(context)
    if result_path:
        doc3.save(result_path + "\Справка " + entryAppNumber.get() + ".docx")
    '''
    if service.get() == 0:
        if physics.get() == 1 and (cultureName[culture.get()] == "озимая пшеница" or cultureName[culture.get()] == "ячмень"):
            doc5 = DocxTemplate(script_path + "\Акт выполненных работ для пшеницы с ГМО и физ-хим (шаблон).docx")
        else:
            doc5 = DocxTemplate(script_path + "\Акт выполненных работ (шаблон).docx")
    else:
        if physics.get() == 1 and (cultureName[culture.get()] == "озимая пшеница" or cultureName[culture.get()] == "ячмень"):
            doc5 = DocxTemplate(script_path + "\Акт выполненных работ для пшеницы с ГМО и физ-хим с услугами (шаблон).docx")
        else:
            doc5 = DocxTemplate(script_path + "\Акт выполненных работ с услугами (шаблон).docx")
    doc5.render(context)
    if result_path:
        doc5.save(result_path + "\Акт выполненных работ " + entryAppNumber.get() + ".docx")
    
    doc6 = openpyxl.load_workbook(script_path + "\Акт отбора (шаблон).xlsx")
    doc6sheet = doc6.active
    i = 0
    for r in range(1,doc6sheet.max_row+1):
        for c in range(1,doc6sheet.max_column+1):
            s = doc6sheet.cell(r,c).value
            for key in context:
                #print(key, '->', context[key])
                temp = '{{ ' + key + ' }}'
                while s != None and re.search(temp, str(s) ,flags=re.I):
                    doc6sheet.cell(r,c).value = re.sub(temp, context[key], str(s), flags=re.I)
                    s = doc6sheet.cell(r,c).value
                    i += 1
    if result_path:
        doc6.save(result_path + "\Акт отбора " + entryAppNumber.get() + ".xlsx")
    '''
    doc6 = DocxTemplate(script_path + "\Акт отбора (шаблон).docx")
    doc6.render(context)
    if result_path:
        doc6.save(result_path + "\Акт отбора " + entryAppNumber.get() + ".docx")
    return context
    
def fail(text):
    messagebox.showinfo("Ошибка", text)
    
def run():
    global result_path
    context = runApp()
    if not context:
        return
    doc1 = DocxTemplate(script_path + "\Договор (шаблон).docx")
    doc1.render(context)
    if result_path:
        doc1.save(result_path + "\Договор.docx")
    doc4 = DocxTemplate(script_path + "\Общая карта клиента (шаблон).docx")
    doc4.render(context)
    if result_path:
        doc4.save(result_path + "\Общая карта клиента.docx")
    return

cultureName = {
    1 : "озимая пшеница",
    2 : "подсолнечник", #масличные
    3 : "ячмень",
    4 : "горох", #зернобобовые
    5 : "кукуруза",
    6 : "просо"
}

cultureCode = {
    "озимая пшеница" : "1001990000",
    "подсолнечник" : "1206009900",
    "ячмень" : "1003900000",
    "горох" : "0713109009",
    "кукуруза" : "1005900000",
    "просо" : "1008290000"
}

selectionGost = {
    "озимая пшеница" : "13586.3 - 2015",
    "подсолнечник" : "10852-86",
    "ячмень" : "13586.3 - 2015",
    "горох" : "13586.3 - 2015", # ???
    "кукуруза" : "13586.3 - 2015",
    "просо" : "13586.3 - 2015"
}

safety = "Комплекс работ по определению показателей безопасности "
secondWord = {
    0 : "(включая ГМО) ",
    1 : "(включая ГМО, физико-химические показатели) ",
    2 : ""
}
keyWord = {
    "озимая пшеница" : "(пшеница, ячмень, и продукты их переработки)",
    "подсолнечник" : "(масличные культуры)",
    "ячмень" : "(пшеница, ячмень, и продукты их переработки)",
    "горох" : "(зернобобовые культуры)",
    "кукуруза" : "(кукуруза, рожь, овёс, рис)",
    "просо" : "(просо, сорго, трикале)"
}

wheatExtra = 582.32
workComplexPrice = {
    "озимая пшеница" : {
        0 : 10826,
        1 : 10826,
        2 : 7887.7
    },
    "подсолнечник" : {
        0 : 6573.3,
        1 : 9006.6,
        2 : 5739.95
    },
    "ячмень" : {
        0 : 10826,
        1 : 10826,
        2 : 7887.7
    },
    "горох" : {
        0 : 8323.3,
        1 : 9156.6,
        2 : 5798.3
    },
    "кукуруза" : {
        0 : 9004.05,
        1 : 9837.35,
        2 : 6545
    },
    "просо" : {
        0 : 9152.2,
        1 : 9985,
        2 : 6610.55
    }
}

startPurpose = "ТР ТС 015/2011 «О безопасности зерна»"
sunflowerPurpose = "ТР ТС 021/2011 «О безопасности пищевой продукции»"
purposePhysics = {
    "озимая пшеница" : "ГОСТ 9353-2016 «Пшеница. Технические условия»",
    "подсолнечник" : "ГОСТ 22391-2015 «Подсолнечник. Технические условия»",
    "ячмень" : "ГОСТ 28672-90 «Ячмень. Требования при заготовках и поставках»",
    "горох" : "ГОСТ 28674-90 «Горох. Требования при заготовках и поставках»",
    "кукуруза" : "ГОСТ 13634-90 «Кукуруза. Требования при заготовках и поставках»",
    "просо" : "ГОСТ 22983-88 «Просо. Требования при заготовках и поставках»"
}
    
root = Tk()
root.resizable(False, False)
root.title("Автозаполнение документов")
root.geometry("1100x900")

main_menu = Menu(root)
file_menu = Menu(main_menu, tearoff=0)

def _onKeyRelease(event):
    ctrl  = (event.state & 0x4) != 0
    if event.keycode==88 and  ctrl and event.keysym.lower() != "x": 
        event.widget.event_generate("<<Cut>>")



    if event.keycode==86 and  ctrl and event.keysym.lower() != "v": 
        event.widget.event_generate("<<Paste>>")



    if event.keycode==67 and  ctrl and event.keysym.lower() != "c":
        event.widget.event_generate("<<Copy>>")

root.bind_all("<Key>", _onKeyRelease, "+")

main_menu.add_cascade(label="Файл", menu=file_menu)
file_menu.add_command(label="Новый файл", command=doWeClear)
file_menu.add_command(label="Открыть", command=choosePath)
file_menu.add_separator()
file_menu.add_command(label="Сохранить", command=save)
file_menu.add_command(label="Сохранить как...", command=saveAs)
main_menu.add_command(label="О программе", command=about)
root.config(menu=main_menu)

make_menu(root)

curX = 15
curY = 10
stepY = 20

labelOrgType = Label(text="Тип организации:", justify=LEFT)
labelOrgType.place(x = curX, y = curY)
orgType = IntVar()
orgType.set(1)

labelOrg = Label(text="Название организации:", justify=LEFT)
labelOrg.place(x = curX + 150, y = curY)

curY = curY + stepY
radioIpkfh = Radiobutton(text="ИП глава К(Ф)Х", value=1, variable=orgType, padx=2, pady=2, command = setOrgType)
radioIpkfh.place(x = curX, y = curY)

messageOrg = StringVar()
entryOrg = Entry(textvariable = messageOrg)
entryOrg.place(x = curX + 150, y = curY, width=300)

curY = curY + stepY
radioOOO = Radiobutton(text="Организация", value=2, variable=orgType, padx=2, pady=2, command = setOrgType)
radioOOO.place(x = curX, y = curY)

labelPost = Label(text="Дожность заявителя:", justify=LEFT)
labelPost.place(x = curX + 150, y = curY)

curY = curY + stepY
messagePost = StringVar()
entryPost = Entry(textvariable = messagePost)
entryPost.place(x = curX + 150, y = curY, width=300)

curY = curY + stepY
labelOwner = Label(text="ФИО заявителя:", justify=LEFT)
labelOwner.place(x = curX, y = curY)

labelDate = Label(text="Дата (формат ДД.ММ):", justify=LEFT)
labelDate.place(x = curX + 350, y = curY)

curY = curY + stepY
messageOwner = StringVar()
entryOwner = Entry(textvariable = messageOwner)
entryOwner.place(x = curX, y = curY, width=300)
entryOwner.bind_class("Entry", "<Button-3><ButtonRelease-3>", show_menu)

messageDate = StringVar()
entryDate = Entry(textvariable = messageDate)
entryDate.place(x = curX + 350, y = curY, width=50)
entryDate.insert(0, day + "." + month)

curY = curY + stepY
labelOwnerInclined = Label(text="ФИО заявителя в родительном падеже:", justify=LEFT)
labelOwnerInclined.place(x = curX, y = curY)

curY = curY + stepY
messageOwnerInclined = StringVar()
entryOwnerInclined = Entry(textvariable = messageOwnerInclined)
entryOwnerInclined.place(x = curX, y = curY, width=300)

curY = curY + stepY
labelBase = Label(text="Документ, подтверждающий полномочия:", justify=LEFT)
labelBase.place(x = curX, y = curY)

curY = curY + stepY
messageBase = StringVar()
entryBase = Entry(textvariable = messageBase)
entryBase.place(x = curX, y = curY, width=500)

curY = curY + stepY
labelResponsiblePerson = Label(text="Ответственное лицо:", justify=LEFT)
labelResponsiblePerson.place(x = curX, y = curY)

curY = curY + stepY
messageResponsiblePerson = StringVar()
entryResponsiblePerson = Entry(textvariable = messageResponsiblePerson)
entryResponsiblePerson.place(x = curX, y = curY, width=300)

curY = curY + stepY
labelResponsiblePerson = Label(text="Документ, подтверждающий его полномочия:", justify=LEFT)
labelResponsiblePerson.place(x = curX, y = curY)

curY = curY + stepY
messageResponsiblePersonBase = StringVar()
entryResponsiblePersonBase = Entry(textvariable = messageResponsiblePersonBase)
entryResponsiblePersonBase.place(x = curX, y = curY, width=500)

curY = curY + stepY
labelResponsiblePersonBaseDate = Label(text="Дата документа подтверждающего его полномочия (формат ДД.ММ.ГГГГ):", justify=LEFT)
labelResponsiblePersonBaseDate.place(x = curX, y = curY)

curY = curY + stepY
messageResponsiblePersonBaseDate = StringVar()
entryResponsiblePersonBaseDate = Entry(textvariable = messageResponsiblePersonBaseDate)
entryResponsiblePersonBaseDate.place(x = curX, y = curY, width=300)

curY = curY + stepY
labelAdressLegal = Label(text="Юридический адрес:", justify=LEFT)
labelAdressLegal.place(x = curX, y = curY)

curY = curY + stepY
messageAdressLegal = StringVar()
entryAdressLegal = Entry(textvariable = messageAdressLegal)
entryAdressLegal.place(x = curX, y = curY, width=500)

curY = curY + stepY
labelAdressFact = Label(text="Фактический адрес:", justify=LEFT)
labelAdressFact.place(x = curX, y = curY)

curY = curY + stepY
messageAdressFact = StringVar()
entryAdressFact = Entry(textvariable = messageAdressFact)
entryAdressFact.place(x = curX, y = curY, width=500)

curY = curY + stepY
labelInn = Label(text="ИНН:", justify=LEFT)
labelInn.place(x = curX, y = curY)

labelInn = Label(text="КПП:", justify=LEFT)
labelInn.place(x = curX + 200, y = curY)

curY = curY + stepY
messageInn = StringVar()
entryInn = Entry(textvariable = messageInn)
entryInn.place(x = curX, y = curY, width=150)

messageKpp = StringVar()
entryKpp = Entry(textvariable = messageKpp)
entryKpp.place(x = curX + 200, y = curY, width=150)

curY = curY + stepY
labelOgrn = Label(text="ОГРН:", justify=LEFT)
labelOgrn.place(x = curX, y = curY)

labelOgrnip = Label(text="ОГРНИП:", justify=LEFT)
labelOgrnip.place(x = curX + 200, y = curY)

curY = curY + stepY
messageOgrn = StringVar()
entryOgrn = Entry(textvariable = messageOgrn)
entryOgrn.place(x = curX, y = curY, width=150)

messageOgrnip = StringVar()
entryOgrnip = Entry(textvariable = messageOgrnip)
entryOgrnip.place(x = curX + 200, y = curY, width=150)

curY = curY + stepY
labelBank = Label(text="Наименование банка:", justify=LEFT)
labelBank.place(x = curX, y = curY)

curY = curY + stepY
messageBank = StringVar()
entryBank = Entry(textvariable = messageBank)
entryBank.place(x = curX, y = curY, width=500)

curY = curY + stepY
labelAccPay = Label(text="Расчетный счет:", justify=LEFT)
labelAccPay.place(x = curX, y = curY)

labelAccCorr = Label(text="Корреспондентский счет:", justify=LEFT)
labelAccCorr.place(x = curX + 250, y = curY)

curY = curY + stepY
messageAccPay = StringVar()
entryAccPay = Entry(textvariable = messageAccPay)
entryAccPay.place(x = curX, y = curY, width=200)

messageAccCorr = StringVar()
entryAccCorr = Entry(textvariable = messageAccCorr)
entryAccCorr.place(x = curX + 250, y = curY, width=200)

curY = curY + stepY
labelBik = Label(text="БИК:", justify=LEFT)
labelBik.place(x = curX, y = curY)

curY = curY + stepY
messageBik = StringVar()
entryBik = Entry(textvariable = messageBik)
entryBik.place(x = curX, y = curY, width=200)

curY = curY + stepY
labelNumber = Label(text="Номер телефона:", justify=LEFT)
labelNumber.place(x = curX, y = curY)

labelMail = Label(text="Почта:", justify=LEFT)
labelMail.place(x = curX + 250, y = curY)

curY = curY + stepY
messageNumber = StringVar()
entryNumber = Entry(textvariable = messageNumber)
entryNumber.place(x = curX, y = curY, width=225)

messageMail = StringVar()
entryMailOrg = Entry(textvariable = messageMail)
entryMailOrg.place(x = curX + 250, y = curY, width=250)

curY = curY + stepY
labelMaker = Label(text="Изготовитель (для КФХ - ФИО):", justify=LEFT)
labelMaker.place(x = curX, y = curY)

curY = curY + stepY
messageMaker = StringVar()
entryMaker = Entry(textvariable = messageMaker)
entryMaker.place(x = curX, y = curY, width=500)

curY = curY + stepY
labelMakerAdressLegal = Label(text="Юридический адрес изготовителя:", justify=LEFT)
labelMakerAdressLegal.place(x = curX, y = curY)

curY = curY + stepY
messageMakerAdressLegal = StringVar()
entryMakerAddressLegal = Entry(textvariable = messageMakerAdressLegal)
entryMakerAddressLegal.place(x = curX, y = curY, width=500)

curY = curY + stepY
labelMakerAdressFact = Label(text="Фактический адрес изготовителя:", justify=LEFT)
labelMakerAdressFact.place(x = curX, y = curY)

curY = curY + stepY
messageMakerAdressFact = StringVar()
entryMakerAddressFact = Entry(textvariable = messageMakerAdressFact)
entryMakerAddressFact.place(x = curX, y = curY, width=500)

curX = 580
curY = 10
labelMail = Label(text="Место отбора:", justify=LEFT)
labelMail.place(x = curX, y = curY)

curY = curY + stepY
messageSelectionPlace = StringVar()
entrySelectionPlace = Entry(textvariable = messageSelectionPlace)
entrySelectionPlace.place(x = curX, y = curY, width=500)

curY = curY + stepY
labelWeight = Label(text="Масса партии:", justify=LEFT)
labelWeight.place(x = curX, y = curY)

labelWeight = Label(text="Год урожая:", justify=LEFT)
labelWeight.place(x = curX + 150, y = curY)

curY = curY + stepY
messageWeight = StringVar()
entryWeight = Entry(textvariable = messageWeight)
entryWeight.place(x = curX, y = curY, width=100)

messageCollectYear = StringVar()
entryCollectYear = Entry(textvariable = messageCollectYear)
entryCollectYear.place(x = curX + 150, y = curY, width=100)
entryCollectYear.insert(0, "20" + year)

curY = curY + stepY
labelWeight = Label(text="Культура:", justify=LEFT)
labelWeight.place(x = curX, y = curY)

culture = IntVar()
culture.set(1)
for i in cultureName:
    curY = curY + stepY
    radioWheat = Radiobutton(text=cultureName[i], value=i, variable=culture, padx=2, pady=2, command = calculateItogo)
    radioWheat.place(x = curX, y = curY)

curY = curY + stepY*2
labelWorkComplex = Label(text="Комплекс работ:", justify=LEFT)
labelWorkComplex.place(x = curX, y = curY)

curY = curY + stepY
gmo = IntVar()
gmo.set(0)
radioGmo = Radiobutton(text="С ГМО", value=0, variable=gmo, padx=2, pady=2, command = calculateItogo)
radioGmo.place(x = curX, y = curY)

physics = IntVar()
physics.set(0)
radioNoPhysics = Radiobutton(text="Без физических показателей", value=0, variable=physics, padx=2, pady=2, command = calculateItogo)
radioNoPhysics.place(x = curX + 100, y = curY)

curY = curY + stepY
radioNoGmo = Radiobutton(text="Без ГМО", value=1, variable=gmo, padx=2, pady=2, command = calculateItogo)
radioNoGmo.place(x = curX, y = curY)

radioPhysics = Radiobutton(text="С физическими показателями", value=1, variable=physics, padx=2, pady=2, command = calculateItogo)
radioPhysics.place(x = curX + 100, y = curY)

curY = curY + stepY*2
service = IntVar()
checkbuttonService = Checkbutton(text = "Консультационные услуги", variable = service, onvalue = 1, offvalue = 0, command = calculateItogo)
checkbuttonService.place(x = curX, y = curY)

curY = curY + stepY*2
labelItogo = Label(text= "Итоговая сумма: " + str2(itogo) + " рублей", justify=LEFT)
calculateItogo()
labelItogo.place(x = curX, y = curY)

curY = curY + stepY*2
labelAppNumber = Label(text="Номер заявления:", justify=LEFT)
labelAppNumber.place(x = curX, y = curY)

labelAppNumber = Label(text="Номер договора:", justify=LEFT)
labelAppNumber.place(x = curX + 150, y = curY)

curY = curY + stepY
messageAppNumber = StringVar()
entryAppNumber = Entry(textvariable = messageAppNumber)
entryAppNumber.place(x = curX, y = curY, width=50)
entryAppNumber.insert(0, "1")

messageContract = StringVar()
entryContract = Entry(textvariable = messageContract)
entryContract.place(x = curX + 150, y = curY, width=50)

curY = curY + stepY*2
buttonRun = Button(text="Без договора", command = runApp)
buttonRun.place(x = curX, y = curY, height=30, width=150)

curY = curY + stepY*2
buttonRun = Button(text="Полный пакет", command = run)
buttonRun.place(x = curX, y = curY, height=30, width=150)

curY = curY + stepY*2
labelHeadFIO = Label(text="ФИО руководителя МРО:", justify=LEFT)
labelHeadFIO.place(x = curX, y = curY)

curY = curY + stepY
headFIO = StringVar()
entryHeadFIO = Entry(textvariable = headFIO)
entryHeadFIO.place(x = curX, y = curY, width=500)
try:
    entryHeadFIO.insert(0, config2['settings']['фио_руководителя'])
except:
    print('Не найдено поле фио_руководителя')
entryHeadFIO['state'] = 'disabled'

curY = curY + stepY
labelHeadFIO = Label(text="ФИО руководителя МРО в родительном падеже:", justify=LEFT)
labelHeadFIO.place(x = curX, y = curY)

curY = curY + stepY
headFIOInclined = StringVar()
entryHeadFIOInclined = Entry(textvariable = headFIOInclined)
entryHeadFIOInclined.place(x = curX, y = curY, width=500)
try:
    entryHeadFIOInclined.insert(0, config2['settings']['фио_руководителя_рп'])
except:
    print('Не найдено поле фио_руководителя_рп')
entryHeadFIOInclined['state'] = 'disabled'

curY = curY + stepY
labelHeadBase = Label(text="На основании чего дейстует руководитель (в родительном падеже):\nпример: Доверенности № 1000 от 01.01.2000", justify=LEFT)
labelHeadBase.place(x = curX, y = curY)
curY = curY + stepY*2
headBase = StringVar()
entryHeadBase = Entry(textvariable = headBase)
entryHeadBase.place(x = curX, y = curY, width=500)
try:
    entryHeadBase.insert(0, config2['settings']['доверенность'])
except:
    print('Не найдено поле доверенность')
entryHeadBase['state'] = 'disabled'

curY = curY + stepY
labelDepId = Label(text="Номер МРО:", justify=LEFT)
labelDepId.place(x = curX, y = curY)

labelDepNameInclined = Label(text="Название МРО в родительном падеже (пример: Азовского):", justify=LEFT)
labelDepNameInclined.place(x = curX + 100, y = curY)

curY = curY + stepY
depId = StringVar()
entryDepId = Entry(textvariable = depId)
entryDepId.place(x = curX, y = curY, width=50)
try:
    entryDepId.insert(0, config2['settings']['номер_отдела'])
except:
    print('Не найдено поле номер_отдела')
entryDepId['state'] = 'disabled'

depNameInclined = StringVar()
entryDepNameInclined = Entry(textvariable = depNameInclined)
entryDepNameInclined.place(x = curX + 100, y = curY, width=300)
try:
    entryDepNameInclined.insert(0, config2['settings']['название_отдела'])
except:
    print('Не найдено поле название_отдела')
entryDepNameInclined['state'] = 'disabled'

curY = curY + stepY
labelMail = Label(text="Почта:", justify=LEFT)
labelMail.place(x = curX, y = curY)
curY = curY + stepY
mail = StringVar()
entryMail = Entry(textvariable = mail)
entryMail.place(x = curX, y = curY, width=500)
try:
    entryMail.insert(0, config2['settings']['почта'])
except:
    print('Не найдено поле почта')
entryMail['state'] = 'disabled'

curY = curY + stepY*2
buttonChangeHeadInfo = Button(text='Изменить информацию о начальнике МРО', command = changeHeadInfo)
buttonChangeHeadInfo.place(x = curX, y = curY, height=40, width=500)

setOrgType()

root.mainloop()